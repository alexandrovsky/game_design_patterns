import random
from flask import Flask
from flask import Response, jsonify, render_template, request
from flask_bootstrap import Bootstrap
import pandas as pd
import openpyxl
app = Flask(__name__)
Bootstrap(app)

WORK_SHEET = 'game_design_patterns.xlsx'

@app.route('/')
def homepage():
    targets = {}
    wb = openpyxl.load_workbook(WORK_SHEET)
    ws = wb.get_sheet_by_name('Sheet1')
    max_row = ws.max_row
    
    for i in range(3):
        row = random.randrange(1, max_row)
        print(row, max_row)
        targets[ws.cell(row=row, column=1).value] = ws.cell(row=row, column=1).hyperlink.target    
    return render_template('index.html', targets=targets)



if __name__ == '__main__':
    app.run(host="0.0.0.0", port=80)